from os import listdir, makedirs
from os.path import isfile, join, exists
from docxtpl import DocxTemplate
from ChangeTracker import change_check
from openpyxl import load_workbook
from shutil import copyfile
from DocxMerge import merge_docs
import datetime


def make_if_not_exist(directory):
    if not exists(directory):
        makedirs(directory)


# return a dictionary generated from an excel spreadsheet passed
# overwrite overlapping entries from result_context
def get_context(workbook_path, result_context):

    wb = load_workbook(workbook_path)

    # for each sheet in the workbook
    for sheet in wb:
        headings = []
        if sheet.title not in result_context:
            result_context[sheet.title] = {}
        first_row = True

        # print(sheet.max_row)
        # if there are only two rows in this sheet,
        # make it sheetname.heading accessible
        if sheet.max_row == 2:

            for col in sheet.columns:  # seems a bit like overkill,
                colm = []  # but it's more readable this way
                for cell in col:
                    if cell.value is not None:
                        colm.append(str(cell.value).strip())  # just doing some housecleaning

                if len(colm) > 1:  # if we have at most two and at least one, woo
                    if colm[1] is not "":
                        result_context[sheet.title][colm[0]] = colm[1]
                        # print(colm[1])
        elif sheet.max_row < 2:  # this means there's just a heading
            continue
        else:

            # otherwise, we make a list of items
            # using the first column as a key
            # also is iterable
            for row in sheet.rows:
                if first_row:
                    first_row = False
                    for cell in row:
                        headings.append(cell.value)
                    continue

                first_cell = ""
                row_index = 0

                for cell in row:
                    text = str(cell.value)

                    # if it's empty, scrap it
                    if cell.value is None or text.strip() is "":
                        row_index += 1
                        continue

                    text = text.strip()
                    if first_cell is "":
                        first_cell = text
                        if first_cell not in result_context[sheet.title]:
                            result_context[sheet.title][first_cell] = {}

                    result_context[sheet.title][first_cell][headings[row_index]] = text
                    row_index += 1

    return result_context


def render_doc(content_context, from_path):
    temp = DocxTemplate(from_path)
    temp.render(content_context)
    return temp.docx


class DocxTemplateManager(object):

    def __init__(self, parent_dir, template_path, completed_path_tpl, archive_path_tpl, default_excel):
        self.parent_dir = parent_dir

        if parent_dir in template_path:
            self.template_path = template_path
            self.template_folder = template_path.replace(parent_dir, "")
        else:
            self.template_path = parent_dir + template_path
            self.template_folder = template_path

        if parent_dir in completed_path_tpl:
            self.completed_path_tpl = completed_path_tpl
            self.completed_folder_tpl = completed_path_tpl.replace(parent_dir, "")
        else:
            self.completed_folder_tpl = completed_path_tpl
            self.completed_path_tpl = parent_dir + completed_path_tpl

        if parent_dir in archive_path_tpl:
            self.archive_path_tpl = archive_path_tpl
            self.archive_folder_tpl = archive_path_tpl.replace(parent_dir, "")
        else:
            self.archive_folder_tpl = archive_path_tpl
            self.archive_path_tpl = parent_dir + archive_path_tpl

        if parent_dir in default_excel:
            self.default_fill = default_excel
            self.default_excel = default_excel.replace(parent_dir, "")
        else:
            self.default_excel = default_excel
            self.default_fill = parent_dir + default_excel

    # using relative file paths
    def build_dependencies(self, item_name, template_names, other_deps=None):

        if other_deps is None:
            other_deps = {}

        all_deps = []
        if "all" in other_deps:
            all_deps = other_deps["all"]
            other_deps.pop("all")

        template_deps = list(all_deps)
        if "template" in other_deps:
            template_deps += other_deps["template"]
            other_deps.pop("template")

        completed_deps = list(all_deps)
        if "completed" in other_deps:
            temp = other_deps["completed"]

            for x in temp:
                t = x
                if "{}" in x:
                    t = x.format(item_name)
                completed_deps.append(t)
            other_deps.pop("completed")

        archive_deps = list(all_deps)
        if "archive" in other_deps:
            temp = other_deps["archive"]

            for x in temp:
                t = x
                if "{}" in x:
                    t = x.format(item_name)
                archive_deps.append(t)

            other_deps.pop("archive")

        completed_folder = self.completed_folder_tpl.format(item_name)
        archive_folder = self.archive_folder_tpl.format(item_name)

        dependency_dict = other_deps
        dependencies = []

        # paths are in relation to the parent_dir
        # every completed file is in relation to it
        for template_name in template_names:
            template_key = self.template_folder + template_name
            dependency_dict[template_key] = {"dependencies": list(template_deps)}
            doc_name = template_name.replace(" Template", "")

            completed_doc_name = completed_folder + doc_name
            dependency_dict[completed_doc_name] = {"dependencies": list(completed_deps)}
            dependency_dict[completed_doc_name]["dependencies"].append(template_key)

            archive_doc_name = archive_folder + doc_name

            dependency_dict[archive_doc_name] = {"dependencies": list(archive_deps)}
            dependency_dict[archive_doc_name]["dependencies"].append(completed_doc_name)

            dependencies.append(completed_doc_name)

        return dependency_dict, dependencies

    # returns true if the document was changed
    def fill_template(self, template_name, in_context, item_dict, completed_name, archive_name):

        completed_path_doc = self.parent_dir + completed_name
        archive_path_doc = self.parent_dir + archive_name

        # if there are changes in dependencies
        if "changes" in item_dict[completed_name] and len(item_dict[completed_name]["changes"]) > 0:
            print("rendering: {}".format(template_name))
            # if there are individual edits in the document aka, the dependency for archived docs, redo changes
            if completed_name in item_dict and isfile(archive_path_doc) and len(item_dict[completed_name]["changes"]) > 0:
                doc = render_doc(in_context, self.template_path + template_name)
                merge_docs(archive_path_doc, completed_path_doc, [doc, completed_path_doc])
            else:
                doc = render_doc(in_context, self.template_path + template_name)
                doc.save(completed_path_doc)

            copyfile(completed_path_doc, archive_path_doc)
            return True

        return False

    def fill_item_templates(self, item_name, in_context, template_names):
        changed_docs = []
        fill_key = item_name.replace("/", "")

        completed_folder = self.completed_folder_tpl.format(fill_key)
        archive_folder = self.archive_folder_tpl.format(fill_key)

        specific_excel = completed_folder + "FillItems.xlsx"
        specific_fill = self.parent_dir + specific_excel

        # print(policies)
        make_if_not_exist(self.parent_dir + completed_folder)
        make_if_not_exist(self.parent_dir + archive_folder)

        if not isfile(specific_fill):
            print("Copying FillItems.xlsx into {}".format(self.parent_dir + completed_folder))
            copyfile(self.default_fill, specific_fill)

        item_dict = {specific_excel: {"dependencies": [self.default_excel]},
                     self.default_excel: {"dependencies": None},
                     "completed": [specific_excel]}

        item_dict, depend = self.build_dependencies(item_name, template_names, item_dict)
        item_dict = change_check(self.parent_dir, item_dict, depend)

        # back to normal files here
        context = get_context(self.default_fill, in_context)  # default items
        context = get_context(specific_fill, context)  # specific items

        _date = datetime.datetime.now()

        context["date"] = {'month': _date.strftime("%B"), 'year': _date.strftime("%Y")}  # add date

        # back to relative path
        for template in template_names:
            filename = template.replace(" Template", "")
            completed_name = completed_folder + filename
            archive_name = archive_folder + filename

            changed = self.fill_template(template, in_context, item_dict, completed_name, archive_name)

            if changed:
                changed_docs.append(filename)

        return changed_docs

    def fill_all_templates(self, in_context, template_names):
        if template_names is None:
            template_names = [f for f in listdir(self.template_path)
                              if isfile(join(self.template_path, f)) and f[0] is not "~"]

        changed_dict = {}
        for key in in_context.keys():
            changed_dict[key] = self.fill_item_templates(item_name=key, in_context=in_context[key],
                                                         template_names=template_names)

        return changed_dict
